from __future__ import annotations

from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.app.services.pdf.structure import collect_page_images, extract_tables, page_reading_items


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 48))
        sheet.column_dimensions[column].width = max(12, max_length + 2)


def _write_table(sheet, rows: list[list[str]], start_row: int = 1) -> int:
    header_fill = PatternFill('solid', fgColor='DCE6F7')
    header_font = Font(bold=True, color='1F2A44')
    thin = Border(
        left=Side(style='thin', color='B7C5DE'),
        right=Side(style='thin', color='B7C5DE'),
        top=Side(style='thin', color='B7C5DE'),
        bottom=Side(style='thin', color='B7C5DE'),
    )

    row_cursor = start_row
    for row_index, row_values in enumerate(rows):
        for col_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_cursor, column=col_index, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin
            if row_index == 0:
                cell.fill = header_fill
                cell.font = header_font
        row_cursor += 1
    return row_cursor


def _lines_to_rows(page: fitz.Page) -> list[list[str]]:
    """Split text lines into columns using x positions when no formal table is detected."""
    rows: list[list[str]] = []
    blocks = page.get_text('dict').get('blocks', [])
    for block in blocks:
        if block.get('type') != 0:
            continue
        for line in block.get('lines', []):
            spans = line.get('spans', [])
            if not spans:
                continue
            # Group spans that are far apart horizontally into separate cells.
            spans_sorted = sorted(spans, key=lambda span: span.get('bbox', [0, 0, 0, 0])[0])
            cells: list[str] = []
            current = spans_sorted[0].get('text', '')
            last_x1 = spans_sorted[0].get('bbox', [0, 0, 0, 0])[2]
            for span in spans_sorted[1:]:
                x0 = span.get('bbox', [0, 0, 0, 0])[0]
                text = span.get('text', '')
                if x0 - last_x1 > 28:
                    cells.append(current.strip())
                    current = text
                else:
                    current += text
                last_x1 = span.get('bbox', [0, 0, 0, 0])[2]
            cells.append(current.strip())
            if any(cells):
                rows.append(cells)
    return rows


def pdf_to_excel(input_path: str | Path, output_path: str | Path) -> str:
    """
    Convert PDF → editable Excel workbook.
    Uses detected tables and structured text rows — no page screenshot sheets.
    """
    source = Path(input_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    with fitz.open(str(source)) as document:
        for page_number, page in enumerate(document, start=1):
            sheet = workbook.create_sheet(f'Page {page_number}')
            tables = extract_tables(page)
            cursor = 1

            if tables:
                for table_index, table in enumerate(tables, start=1):
                    if table_index > 1:
                        cursor += 1
                    sheet.cell(row=cursor, column=1, value=f'Table {table_index}').font = Font(bold=True, color='4134EA')
                    cursor += 1
                    cursor = _write_table(sheet, table.rows, start_row=cursor)
                    cursor += 1
            else:
                rows = _lines_to_rows(page)
                if rows:
                    cursor = _write_table(sheet, rows, start_row=1)
                else:
                    text = (page.get_text('text') or '').strip()
                    if text:
                        for row_number, line in enumerate(text.splitlines(), start=1):
                            sheet.cell(row=row_number, column=1, value=line)
                    else:
                        sheet.cell(row=1, column=1, value='[No extractable table/text on this page]')

            # Also capture non-table paragraph text on a second area if tables existed,
            # without duplicating table content.
            if tables:
                extras: list[str] = []
                for kind, payload in page_reading_items(document, page):
                    if kind != 'block' or payload.kind != 'text':
                        continue
                    paragraph = ' '.join(
                        ''.join(span.text for span in line.spans).strip()
                        for line in payload.lines
                    ).strip()
                    if paragraph:
                        extras.append(paragraph)
                if extras:
                    cursor += 1
                    sheet.cell(row=cursor, column=1, value='Additional text').font = Font(bold=True, color='4134EA')
                    cursor += 1
                    for line in extras:
                        sheet.cell(row=cursor, column=1, value=line)
                        cursor += 1

            # Embed real PDF images (not full-page screenshots) under the data.
            page_images = collect_page_images(document, page)
            if page_images:
                cursor += 2
                sheet.cell(row=cursor, column=1, value='Images').font = Font(bold=True, color='4134EA')
                cursor += 1
                for image_index, image_block in enumerate(page_images, start=1):
                    if not image_block.image_bytes:
                        continue
                    image_path = output.with_name(f'.page-{page_number}-img-{image_index}.{image_block.image_ext or "png"}')
                    image_path.write_bytes(image_block.image_bytes)
                    try:
                        excel_image = SpreadsheetImage(str(image_path))
                        excel_image.anchor = f'A{cursor}'
                        # Keep images readable but not gigantic.
                        excel_image.width = min(getattr(excel_image, 'width', 320) or 320, 420)
                        excel_image.height = min(getattr(excel_image, 'height', 240) or 240, 320)
                        sheet.add_image(excel_image)
                        cursor += 18
                    except Exception:
                        sheet.cell(row=cursor, column=1, value=f'[Image {image_index} could not be embedded]')
                        cursor += 1

            _autosize(sheet)
            sheet.freeze_panes = 'A2'

    if not workbook.sheetnames:
        workbook.create_sheet('Page 1')
    workbook.save(output)
    for snapshot_path in output.parent.glob('.page-*-img-*.*'):
        snapshot_path.unlink(missing_ok=True)
    return str(output)
