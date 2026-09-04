from __future__ import annotations

from html import escape
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .office_renderer import has_libreoffice, render_office_to_pdf


def _fallback_excel_to_pdf(source: Path, output: Path) -> str:
    workbook = load_workbook(source, read_only=True, data_only=True)
    styles = getSampleStyleSheet()
    story = []

    for sheet_index, sheet in enumerate(workbook.worksheets):
        if sheet_index:
            story.append(Spacer(1, 18))
        story.append(Paragraph(escape(sheet.title), styles['Heading2']))
        story.append(Spacer(1, 8))

        rows = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= 80:
                rows.append([Paragraph('… truncated …', styles['Normal'])])
                break
            cells = [Paragraph(escape('' if value is None else str(value)), styles['Normal']) for value in row[:12]]
            if cells:
                rows.append(cells)

        if not rows:
            story.append(Paragraph('Sheet is empty.', styles['Normal']))
            continue

        column_count = max(len(row) for row in rows)
        col_width = min(1.4 * inch, (10 * inch) / max(column_count, 1))
        table = Table(rows, colWidths=[col_width] * column_count, hAlign='LEFT', repeatRows=1)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.4, colors.Color(0.7, 0.75, 0.82)),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.93, 0.95, 0.99)),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(table)

    workbook.close()
    pdf = SimpleDocTemplate(
        str(output),
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    pdf.build(story or [Paragraph('Converted spreadsheet is empty.', styles['Normal'])])
    return str(output)


def excel_to_pdf(input_path: str | Path, output_path: str | Path) -> str:
    source = Path(input_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    if has_libreoffice():
        rendered = render_office_to_pdf(source, output)
        if rendered:
            return rendered
        rendered = render_office_to_pdf(source, output)
        if rendered:
            return rendered
        raise RuntimeError(
            'LibreOffice is installed but could not convert this Excel file. '
            'Close other LibreOffice windows and try again.'
        )

    return _fallback_excel_to_pdf(source, output)
